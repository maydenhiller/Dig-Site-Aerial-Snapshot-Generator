# app.py
import streamlit as st
import io
import zipfile
import traceback
import requests
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook

st.set_page_config(page_title="Dig Site Aerial Snapshot Generator", layout="centered")
st.title("📑 Dig Site Aerial Snapshot Generator (robust with debugger)")

# --- Debug logger (always renders) ---
if "debug_log" not in st.session_state:
    st.session_state.debug_log = []

def log(msg):
    s = str(msg)
    st.session_state.debug_log.append(s)
    st.write(s)

def log_exception(ctx, exc):
    tb = traceback.format_exc()
    st.session_state.debug_log.append(f"[ERROR] {ctx}: {exc}\n{tb}")
    st.error(f"{ctx}: {exc}")
    with st.expander("Debugger — full traceback"):
        st.code(tb, language="text")

# --- Mapbox token handling (never crashes if missing) ---
def get_token():
    try:
        token = st.secrets["mapbox"]["token"]
        log("Loaded Mapbox token from st.secrets['mapbox']['token'].")
        return token
    except Exception as e:
        log(f"No Mapbox token in secrets: {e}")
        return None

MAPBOX_TOKEN = get_token()
if not MAPBOX_TOKEN:
    MAPBOX_TOKEN = st.sidebar.text_input("Mapbox access token", type="password", help="Enter your Mapbox token if not set in secrets.")
    if MAPBOX_TOKEN:
        log("Using Mapbox token provided via sidebar input.")
    else:
        st.sidebar.warning("Mapbox token is required. Add to secrets as [mapbox]['token'] or paste here.")

# --- UI: file upload (always visible) ---
uploaded_file = st.file_uploader("Upload Excel file (.xlsx or .xlsm)", type=["xlsx", "xlsm"])

# --- Mapbox fetch + annotate (with defensive coding) ---
def fetch_satellite_image(lat, lon, label, token):
    width_px = 640   # 6.69" wide
    height_px = 312  # 3.25" high
    zoom = 18
    bearing = 0

    url = (
        f"https://api.mapbox.com/styles/v1/mapbox/satellite-v9/static/"
        f"{lon},{lat},{zoom},{bearing}/{width_px}x{height_px}?access_token={token}"
    )
    try:
        resp = requests.get(url, timeout=20)
    except Exception as e:
        log(f"[Mapbox] Request error for '{label}': {e}")
        return None
    if resp.status_code != 200:
        log(f"[Mapbox] Non-200 for '{label}': status={resp.status_code}, body={resp.text[:200]}")
        return None

    try:
        image = Image.open(io.BytesIO(resp.content)).convert("RGB")
    except Exception as e:
        log(f"[PIL] Unable to open image for '{label}': {e}")
        return None

    draw = ImageDraw.Draw(image)

    # Font setup
    font_size = 28
    try:
        font = ImageFont.truetype("arial.ttf", font_size)
    except Exception as e:
        log(f"[Font] Arial not found, fallback to default: {e}")
        font = ImageFont.load_default()

    label = (label or "").upper().strip()

    # Center point
    center_x = image.width // 2
    center_y = image.height // 2

    # Yellow dot at center
    dot_radius = 6
    draw.ellipse(
        [
            (center_x - dot_radius, center_y - dot_radius),
            (center_x + dot_radius, center_y + dot_radius)
        ],
        fill="yellow",
        outline="black"
    )

    # Measure text size
    try:
        bbox = font.getbbox(label)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]
    except Exception as e:
        log(f"[Font] getbbox failed, using textlength: {e}")
        try:
            text_width = draw.textlength(label, font=font)
            text_height = font_size
        except Exception as e2:
            log(f"[Font] textlength failed, using defaults: {e2}")
            text_width, text_height = 100, font_size

    # Position label above/right of dot, keep in bounds
    label_x = max(6, min(center_x + dot_radius + 6, image.width - text_width - 12))
    label_y = max(6, min(center_y - text_height - 6, image.height - text_height - 12))

    # White background with black border
    box_margin = 6
    box_coords = [
        label_x - box_margin,
        label_y - box_margin,
        label_x + text_width + box_margin,
        label_y + text_height + box_margin
    ]
    draw.rectangle(box_coords, fill="white", outline="black")

    # Draw text
    draw.text((label_x, label_y), label, fill="black", font=font)

    buffer = io.BytesIO()
    image.save(buffer, format="JPEG")
    buffer.seek(0)
    return buffer.read()

# --- Main (defensive flow; debugger always visible) ---
try:
    # Always show debugger contents, even before running
    with st.expander("Debugger — activity log"):
        st.code("\n".join(st.session_state.debug_log) or "(no logs yet)", language="text")

    if uploaded_file:
        log(f"Received file: name={getattr(uploaded_file, 'name', 'unknown')}, type={getattr(uploaded_file, 'type', 'unknown')}")

        if not MAPBOX_TOKEN:
            st.error("No Mapbox token provided. Add it in secrets or the sidebar to proceed.")
        else:
            # Robust workbook load
            try:
                uploaded_file.seek(0)
                data = uploaded_file.read()
                log(f"Read uploaded file bytes: {len(data)}")
                wb = load_workbook(io.BytesIO(data), data_only=True)
                log(f"Workbook loaded. Sheets: {wb.sheetnames}")
            except Exception as e:
                log_exception("Failed to open workbook", e)
                wb = None

            if wb:
                # Filter sheets: start with "dig" and not exactly "dig list"
                dig_tabs = [
                    s for s in wb.sheetnames
                    if s.lower().startswith("dig") and s.lower() != "dig list"
                ]
                log(f"Filtered Dig tabs: {dig_tabs}")

                if not dig_tabs:
                    st.error("No valid Dig tabs found (excluding 'Dig list').")
                else:
                    if st.button("Generate Aerial Images"):
                        zip_buffer = io.BytesIO()
                        images_written = 0

                        with zipfile.ZipFile(zip_buffer, "w") as zip_file:
                            for sheet in dig_tabs:
                                try:
                                    ws = wb[sheet]

                                    # Read AR15 and AS15 directly (bypass formulas in J13/J14)
                                    lat_val = ws["AR15"].value  # Latitude
                                    lon_val = ws["AS15"].value  # Longitude
                                    log(f"{sheet}: AR15={lat_val}, AS15={lon_val}")

                                    if lat_val is None or lon_val is None:
                                        st.warning(f"Skipping {sheet}: AR15/AS15 are empty")
                                        log(f"Skipping {sheet}: AR15/AS15 are empty")
                                        continue

                                    # Normalize and convert
                                    try:
                                        lat = float(str(lat_val).strip())
                                        lon = float(str(lon_val).strip())
                                    except Exception as e:
                                        st.warning(f"Skipping {sheet}: could not convert AR15/AS15 to float ({e})")
                                        log(f"Skipping {sheet}: float conversion failed: {e}")
                                        continue

                                    # Bounds check
                                    if not (-90 <= lat <= 90 and -180 <= lon <= 180):
                                        st.warning(f"Skipping {sheet}: coordinates out of bounds (lat {lat}, lon {lon})")
                                        log(f"Skipping {sheet}: out-of-bounds ({lat}, {lon})")
                                        continue

                                    img = fetch_satellite_image(lat, lon, sheet, MAPBOX_TOKEN)
                                    if img:
                                        zip_file.writestr(f"{sheet}.jpg", img)
                                        images_written += 1
                                        log(f"Image written for {sheet}")
                                    else:
                                        st.warning(f"Skipping {sheet}: Mapbox image fetch failed")
                                        log(f"Mapbox image fetch failed for {sheet}")

                                except Exception as e:
                                    log_exception(f"Unhandled error processing sheet '{sheet}'", e)
                                    continue

                        if images_written > 0:
                            zip_buffer.seek(0)
                            st.download_button(
                                label=f"📦 Download {images_written} Dig Site Images ZIP",
                                data=zip_buffer,
                                file_name="dig_site_images.zip",
                                mime="application/zip"
                            )
                            log(f"Prepared ZIP with {images_written} images")
                        else:
                            st.error("No images generated. Check AR15/AS15 values and Mapbox token.")
                            log("No images generated at end of run")
except Exception as e:
    log_exception("Top-level app crash", e)

# --- Attribution footer ---
st.caption("© Mapbox © OpenStreetMap contributors")
